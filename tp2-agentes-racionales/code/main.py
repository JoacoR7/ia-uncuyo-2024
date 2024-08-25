from environment import *
from reflexiveAgent import *
from randomAgent import *
import copy
import openpyxl

sizes = [2, 4, 8, 16, 32, 64, 128]
dirt_rate = [0.1, 0.2, 0.4, 0.8]
environments = []

for i in sizes:
    for j in dirt_rate:
        for k in range(0, 10):
            environments.append(Environment(i, i, j))

#Se crea el libro excel
excelBook = openpyxl.Workbook()
#Se selecciona la hoja activa
sheet = excelBook.active
#Se colocan las etiquetas en la primera columna
sheet.cell(row = 1, column = 1, value = "Entorno (Tamaño | Dirt Rate)")
sheet.cell(row = 2, column = 1, value = "Cantidad de celdas sucias")
sheet.cell(row = 3, column = 1, value = "Performance Simple")
sheet.cell(row = 4, column = 1, value = "Movimientos Simple")
sheet.cell(row = 5, column = 1, value = "Performance Aleatorio")
sheet.cell(row = 6, column = 1, value = "Movimientos Aleatorio")

n = len(environments)
for i in range(2, n + 2):
    #Se realiza una copia del environment para que el mismo entorno sea utilizado por ambos agentes.
    reflexiveEnvironment = environments[i-2]
    randomEnvironment = copy.deepcopy(reflexiveEnvironment)
    agent1 = ReflexiveAgent(reflexiveEnvironment)
    agent2 = RandomAgent(randomEnvironment)
    sheet.cell(row = 1, column = i, value = "(" + str(reflexiveEnvironment.sizeX) + " x " + str(reflexiveEnvironment.sizeY) + " | " + str(reflexiveEnvironment.dirt_rate) + ")")
    sheet.cell(row = 2, column = i, value = reflexiveEnvironment.initial_dirty_cells_count)
    performance, movimientosMax = reflexiveEnvironment.agent_action(agent1, 1000)
    sheet.cell(row = 3, column = i, value = performance)
    sheet.cell(row = 4, column = i, value = movimientosMax)
    performance, movimientosMax = randomEnvironment.agent_action(agent2, 1000)
    sheet.cell(row = 5, column = i, value = performance)
    sheet.cell(row = 6, column = i, value = movimientosMax)




