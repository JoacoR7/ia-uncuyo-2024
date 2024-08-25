from environment import *
from reflexiveAgent import *
from randomAgent import *
import copy
import openpyxl

sizeList = [2, 4, 8, 16, 32, 64, 128]
dirt_rate_list = [0.1, 0.2, 0.4, 0.8]
environments = []

for i in sizeList:
    for j in dirt_rate_list:
        for k in range(0, 10):
            environments.append(Environment(i, i, j))

#Se crea el libro excel
performanceData = openpyxl.Workbook()

for size in range(len(sizeList)):
    #Nombrado de hojas por entorno: (2x2), (4x4), ..., (128x128)
    sheet_name = f"{2**(size+1)}x{2**(size+1)}"
    performanceDataSheet = performanceData.create_sheet(title=sheet_name)
    performanceDataSheet.cell(row = 1, column = 1, value = "Dirt Rate")
    performanceDataSheet.cell(row = 2, column = 1, value = "Celdas sucias")
    performanceDataSheet.cell(row = 3, column = 1, value = "Max Reflexivo Simple")
    performanceDataSheet.cell(row = 4, column = 1, value = "Max Random")
    reflexiveEnvironments = []
    randomEnvironments = []
    for rate in range(len(dirt_rate_list)):
        #Nombrado de columnas por dirt rate, 10 de cada tasa
        for k in range(1,11):
            performanceDataSheet.cell(row = 1, column = k+1 + (10*rate), value = dirt_rate_list[rate])
        
        max_performance = 0
        #Se pidieron 10 combinaciones de cada entorno y dirt rate
        for i in range(10):
            #Así obtengo los 40 entornos de cada tamaño
            environmentNumber = (size*40) + i + (rate*10)
            reflexiveEnvironment = environments[environmentNumber]
            reflexiveEnvironments.append(reflexiveEnvironment)
            randomEnvironment = copy.deepcopy(reflexiveEnvironment)
            randomEnvironments.append(randomEnvironment)
            agent1 = ReflexiveAgent(reflexiveEnvironment)
            agent2 = RandomAgent(randomEnvironment)

            reflexiveEnvironment.agent_action(agent1, 1000)
            randomEnvironment.agent_action(agent2, 1000)

            #Cantidad de celdas sucias
            performanceDataSheet.cell(row = 2, column = i + 2 + (10*rate), value = reflexiveEnvironment.initial_dirty_cells_count)

            #Performance máximo obtenido
            performanceDataSheet.cell(row=3, column = i + 2 + (10*rate), value = reflexiveEnvironment.performance)
            performanceDataSheet.cell(row=4, column = i + 2 + (10*rate), value = randomEnvironment.performance)

            _, max_performance_aux = max_performance_points([reflexiveEnvironment, randomEnvironment])

            if max_performance < max_performance_aux:
                max_performance = max_performance_aux
    
    #Puntajes de performance por entorno
    for i in range(max_performance):
        performanceDataSheet.cell(row = 5+i, column = 1, value = f"{i+1} SRPA")
        performanceDataSheet.cell(row = 5+i+max_performance, column = 1, value = f"{i+1} RPA")

    for rate in range(len(dirt_rate_list)):
        for i in range(10):
            environmentNumber = i + (10*rate)
            for j in range(max_performance):
                columnNumber = 2 + i + (10*rate)
                if j < len(reflexiveEnvironments[environmentNumber].action_count):
                    performanceDataSheet.cell(row = j + 5, column = columnNumber, value = reflexiveEnvironments[environmentNumber].action_count[j])
                else:
                    performanceDataSheet.cell(row = j + 5, column = columnNumber, value = "-")
                if j < len(randomEnvironments[environmentNumber].action_count):
                    performanceDataSheet.cell(row = j + 5 + max_performance, column = columnNumber, value = randomEnvironments[environmentNumber].action_count[j])
                else:
                    performanceDataSheet.cell(row = j + 5 + max_performance, column = columnNumber, value = "-")
        

if "Sheet" in performanceData.sheetnames:
    del performanceData["Sheet"]
performanceData.save(r'C:\Users\joaqu\Desktop\pruebas.xlsx')



